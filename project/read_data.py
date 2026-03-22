"""
Reads Kapital.xlsx and exports structured JSON for the Steuertool web app.

Output: data/data.json
"""

import json
import openpyxl
from pathlib import Path

XLSX_PATH = Path(__file__).parent / "data" / "Kapital.xlsx"
JSON_PATH = Path(__file__).parent / "data" / "data.json"

INCOME_KEYS = ["lohn", "gewerbe_vv", "kap_ertrag", "erbschaften", "transfers"]
TAX_KEYS    = ["mwst", "est", "kv_pv", "kap_st", "erbe_st", "sonst"]

INCOME_LABELS = {
    "lohn":       "Lohn / Gehalt",
    "gewerbe_vv": "Gewerbe / V&V",
    "kap_ertrag": "Kapitalertrag",
    "erbschaften":"Erbschaften",
    "transfers":  "Transfers",
}
TAX_LABELS = {
    "mwst":    "Umsatzsteuer",
    "est":     "Einkommensteuer",
    "kv_pv":   "Sozialabgaben (KV/PV)",
    "kap_st":  "Kapitalertragsteuer",
    "erbe_st": "Erbschaftsteuer",
    "sonst":   "Sonstiges",
}


def parse_haupttabelle(ws):
    """Parse the 10-group summary table (D1–D9, 91–99%, Top 1%)."""
    groups = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        label = row[1]
        if label is None or not isinstance(label, str):
            continue
        if label.strip().lower().startswith("tool"):
            break

        zufluss = row[2]
        income_values  = [row[3], row[4], row[5], row[6], row[7]]
        abgaben = row[8]
        tax_values     = [row[9], row[10], row[11], row[12], row[13], row[14]]

        # Skip totals row
        if label.lower() == "gesamt":
            continue

        entry = {
            "gruppe": label.strip(),
            "zufluss": zufluss,
            "abgaben": abgaben,
            "einkommen": {k: v for k, v in zip(INCOME_KEYS, income_values)},
            "steuern":   {k: v for k, v in zip(TAX_KEYS,    tax_values)},
        }
        groups.append(entry)

    return groups


def parse_perzentile(ws):
    """Parse the 5%-percentile breakdown (20 bands from 0–5% to 95–100%)."""
    groups = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        label = row[1]
        if label is None or not isinstance(label, str):
            continue
        if label.strip().lower().startswith(("tool", "gesamt")):
            continue
        if not any(c.isdigit() for c in label):
            continue

        income_values = [row[3], row[4], row[5], row[6], row[7]]
        tax_values    = [row[9], row[10], row[11], row[12], row[13], row[14]]

        # Compute sums (some cells contain Excel formula strings in openpyxl)
        zufluss = sum(v for v in income_values if isinstance(v, (int, float)))
        abgaben = sum(v for v in tax_values    if isinstance(v, (int, float)))

        entry = {
            "gruppe": label.strip(),
            "zufluss": round(zufluss, 4),
            "abgaben": round(abgaben, 4),
            "einkommen": {k: (v if isinstance(v, (int, float)) else None)
                          for k, v in zip(INCOME_KEYS, income_values)},
            "steuern":   {k: (v if isinstance(v, (int, float)) else None)
                          for k, v in zip(TAX_KEYS,    tax_values)},
        }
        groups.append(entry)

    return groups


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    haupttabelle = parse_haupttabelle(wb["Haupttabelle"])
    perzentile   = parse_perzentile(wb["5%-Perzentile"])

    output = {
        "meta": {
            "einheit": "Mrd. EUR / Jahr",
            "quelle":  "Kapital.xlsx",
            "labels": {
                "einkommen": INCOME_LABELS,
                "steuern":   TAX_LABELS,
            },
        },
        "haupttabelle": haupttabelle,
        "perzentile":   perzentile,
    }

    JSON_PATH.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Written {len(haupttabelle)} groups (Haupttabelle) + "
          f"{len(perzentile)} bands (5%-Perzentile) -> {JSON_PATH}")


if __name__ == "__main__":
    main()
